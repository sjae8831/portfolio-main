"""
QC Management - PythonAnywhere 웹 배포용
"""
import os
import json
import uuid as uuid_lib
from datetime import datetime

from flask import Flask, request, jsonify, send_from_directory, send_file
import sqlite3

# ─── 경로 설정 (PythonAnywhere용) ───────────────────────
# ★ 아래 USERNAME을 본인 PythonAnywhere 유저네임으로 변경하세요 ★
USERNAME = 'BKCQC'

BASE_DIR = f'/home/{USERNAME}/qc_app'
DB_PATH = os.path.join(BASE_DIR, 'qc_data.db')
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
STATIC_DIR = os.path.join(BASE_DIR, 'static')

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(STATIC_DIR, exist_ok=True)

# 허용 확장자
ALLOWED_EXTENSIONS = {
    'image': {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'},
    'document': {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.csv', '.hwp'},
    'archive': {'.zip', '.rar', '.7z'},
}
ALL_ALLOWED = set()
for exts in ALLOWED_EXTENSIONS.values():
    ALL_ALLOWED.update(exts)

MAX_FILE_SIZE = 50 * 1024 * 1024

# ─── Flask 앱 ───────────────────────────────────
app = Flask(__name__, static_folder=STATIC_DIR)
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# ─── DB 초기화 ───────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS defects (
        id TEXT PRIMARY KEY,
        production_date TEXT DEFAULT '',
        date TEXT NOT NULL,
        product TEXT NOT NULL,
        spec TEXT DEFAULT '',
        lot_no TEXT DEFAULT '',
        line TEXT DEFAULT '',
        worker TEXT DEFAULT '',
        defect_type TEXT NOT NULL,
        quantity TEXT DEFAULT '',
        status TEXT DEFAULT '대기',
        priority TEXT DEFAULT 'medium',
        cause TEXT DEFAULT '',
        action TEXT DEFAULT '',
        responsible TEXT DEFAULT '',
        due_date TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        history TEXT DEFAULT '[]',
        created_at TEXT,
        updated_at TEXT
    )''')
    try: conn.execute("ALTER TABLE defects ADD COLUMN production_date TEXT DEFAULT ''")
    except: pass
    try: conn.execute("ALTER TABLE defects ADD COLUMN spec TEXT DEFAULT ''")
    except: pass
    try: conn.execute("ALTER TABLE defects ADD COLUMN worker TEXT DEFAULT ''")
    except: pass

    c.execute('''CREATE TABLE IF NOT EXISTS tests (
        id TEXT PRIMARY KEY,
        request_date TEXT NOT NULL,
        requester TEXT NOT NULL,
        product TEXT NOT NULL,
        lot_no TEXT DEFAULT '',
        test_type TEXT NOT NULL,
        sample_info TEXT DEFAULT '',
        test_condition TEXT DEFAULT '',
        standard TEXT DEFAULT '',
        status TEXT DEFAULT '대기',
        priority TEXT DEFAULT 'medium',
        due_date TEXT DEFAULT '',
        result TEXT DEFAULT '',
        judgment TEXT DEFAULT '',
        tester TEXT DEFAULT '',
        completed_date TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        created_at TEXT,
        updated_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS certs (
        id TEXT PRIMARY KEY,
        date TEXT NOT NULL,
        cert_no TEXT DEFAULT '',
        product TEXT NOT NULL,
        lot_no TEXT DEFAULT '',
        customer TEXT DEFAULT '',
        category TEXT NOT NULL,
        issue TEXT NOT NULL,
        impact TEXT DEFAULT '',
        action TEXT DEFAULT '',
        status TEXT DEFAULT '대기',
        responsible TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        created_at TEXT,
        updated_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS requests (
        id TEXT PRIMARY KEY,
        request_date TEXT NOT NULL,
        request_type TEXT NOT NULL,
        department TEXT DEFAULT '',
        requester TEXT NOT NULL,
        product TEXT DEFAULT '',
        customer TEXT DEFAULT '',
        detail TEXT NOT NULL,
        status TEXT DEFAULT '대기',
        priority TEXT DEFAULT 'medium',
        due_date TEXT DEFAULT '',
        completed_date TEXT DEFAULT '',
        response TEXT DEFAULT '',
        responsible TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        created_at TEXT,
        updated_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS attachments (
        id TEXT PRIMARY KEY,
        parent_table TEXT NOT NULL,
        parent_id TEXT NOT NULL,
        original_name TEXT NOT NULL,
        stored_name TEXT NOT NULL,
        file_size INTEGER DEFAULT 0,
        file_type TEXT DEFAULT '',
        mime_type TEXT DEFAULT '',
        uploaded_at TEXT NOT NULL
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS issues (
        id TEXT PRIMARY KEY,
        date TEXT NOT NULL,
        category TEXT NOT NULL,
        title TEXT NOT NULL,
        target_line TEXT DEFAULT '',
        target_product TEXT DEFAULT '',
        detail TEXT NOT NULL,
        cause TEXT DEFAULT '',
        impact TEXT DEFAULT '',
        action_taken TEXT DEFAULT '',
        action_plan TEXT DEFAULT '',
        status TEXT DEFAULT '대기',
        priority TEXT DEFAULT 'medium',
        reporter TEXT DEFAULT '',
        responsible TEXT DEFAULT '',
        due_date TEXT DEFAULT '',
        completed_date TEXT DEFAULT '',
        related_dept TEXT DEFAULT '',
        notify_target TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        history TEXT DEFAULT '[]',
        created_at TEXT,
        updated_at TEXT
    )''')

    # 회의록 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS meetings (
        id TEXT PRIMARY KEY,
        date TEXT NOT NULL,
        start_time TEXT DEFAULT '',
        end_time TEXT DEFAULT '',
        title TEXT NOT NULL,
        meeting_type TEXT DEFAULT '',
        location TEXT DEFAULT '',
        organizer TEXT DEFAULT '',
        attendees TEXT DEFAULT '',
        agenda TEXT DEFAULT '',
        content TEXT NOT NULL,
        decisions TEXT DEFAULT '',
        action_items TEXT DEFAULT '',
        next_meeting TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        created_at TEXT,
        updated_at TEXT
    )''')

    # 반품현황 테이블
    c.execute('''CREATE TABLE IF NOT EXISTS returns (
        id TEXT PRIMARY KEY,
        return_date TEXT NOT NULL,
        customer TEXT NOT NULL,
        product TEXT NOT NULL,
        spec TEXT DEFAULT '',
        lot_no TEXT DEFAULT '',
        delivery_date TEXT DEFAULT '',
        return_reason TEXT NOT NULL,
        quantity TEXT DEFAULT '',
        claim_detail TEXT DEFAULT '',
        inspection_result TEXT DEFAULT '',
        cause TEXT DEFAULT '',
        action TEXT DEFAULT '',
        disposition TEXT DEFAULT '',
        status TEXT DEFAULT '대기',
        priority TEXT DEFAULT 'medium',
        responsible TEXT DEFAULT '',
        sales_rep TEXT DEFAULT '',
        due_date TEXT DEFAULT '',
        completed_date TEXT DEFAULT '',
        cost TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        created_at TEXT,
        updated_at TEXT
    )''')

    conn.commit()
    conn.close()

# DB 초기화 실행
init_db()

# ─── 유틸리티 ───────────────────────────────────
def now_str():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def row_to_dict(row):
    if row is None: return None
    return dict(row)

def rows_to_list(rows):
    return [dict(r) for r in rows]

def get_file_category(ext):
    ext = ext.lower()
    for cat, exts in ALLOWED_EXTENSIONS.items():
        if ext in exts: return cat
    return 'other'

# ─── 공통 CRUD ───────────────────────────────────
TABLES = {
    'defects': {
        'fields': ['id','production_date','date','product','spec','lot_no','line','worker',
                    'defect_type','quantity','status','priority','cause','action',
                    'responsible','due_date','notes','history','created_at','updated_at'],
    },
    'tests': {
        'fields': ['id','request_date','requester','product','lot_no','test_type',
                    'sample_info','test_condition','standard','status','priority',
                    'due_date','result','judgment','tester','completed_date','notes',
                    'created_at','updated_at'],
    },
    'certs': {
        'fields': ['id','date','cert_no','product','lot_no','customer','category',
                    'issue','impact','action','status','responsible','notes',
                    'created_at','updated_at'],
    },
    'requests': {
        'fields': ['id','request_date','request_type','department','requester',
                    'product','customer','detail','status','priority','due_date',
                    'completed_date','response','responsible','notes',
                    'created_at','updated_at'],
    },
    'issues': {
        'fields': ['id','date','category','title','target_line','target_product',
                    'detail','cause','impact','action_taken','action_plan','status',
                    'priority','reporter','responsible','due_date','completed_date',
                    'related_dept','notify_target','notes','history',
                    'created_at','updated_at'],
    },
    'meetings': {
        'fields': ['id','date','start_time','end_time','title','meeting_type',
                    'location','organizer','attendees','agenda','content',
                    'decisions','action_items','next_meeting','notes',
                    'created_at','updated_at'],
    },
    'returns': {
        'fields': ['id','return_date','customer','product','spec','lot_no',
                    'delivery_date','return_reason','quantity','claim_detail',
                    'inspection_result','cause','action','disposition','status',
                    'priority','responsible','sales_rep','due_date',
                    'completed_date','cost','notes','created_at','updated_at'],
    },
}

# ─── API 라우트 ───────────────────────────────────
@app.route('/')
def index():
    return send_from_directory(STATIC_DIR, 'index.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(STATIC_DIR, filename)

@app.route('/api/<table>', methods=['GET'])
def get_all(table):
    if table not in TABLES: return jsonify({'error': 'Invalid table'}), 400
    conn = get_db()
    rows = conn.execute(f'SELECT * FROM {table} ORDER BY created_at DESC').fetchall()
    result = rows_to_list(rows)
    for item in result:
        count = conn.execute('SELECT COUNT(*) as cnt FROM attachments WHERE parent_table=? AND parent_id=?', (table, item['id'])).fetchone()
        item['attachment_count'] = count['cnt'] if count else 0
    conn.close()
    return jsonify(result)

@app.route('/api/<table>/<item_id>', methods=['GET'])
def get_one(table, item_id):
    if table not in TABLES: return jsonify({'error': 'Invalid table'}), 400
    conn = get_db()
    row = conn.execute(f'SELECT * FROM {table} WHERE id = ?', (item_id,)).fetchone()
    conn.close()
    if row is None: return jsonify({'error': 'Not found'}), 404
    return jsonify(row_to_dict(row))

@app.route('/api/<table>', methods=['POST'])
def create(table):
    if table not in TABLES: return jsonify({'error': 'Invalid table'}), 400
    data = request.json
    fields = TABLES[table]['fields']
    data['created_at'] = now_str()
    data['updated_at'] = now_str()
    if 'history' in data and isinstance(data['history'], list):
        data['history'] = json.dumps(data['history'], ensure_ascii=False)
    cols = [f for f in fields if f in data]
    vals = [data[f] for f in cols]
    placeholders = ','.join(['?' for _ in cols])
    col_names = ','.join(cols)
    conn = get_db()
    conn.execute(f'INSERT INTO {table} ({col_names}) VALUES ({placeholders})', vals)
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': data.get('id')})

@app.route('/api/<table>/<item_id>', methods=['PUT'])
def update(table, item_id):
    if table not in TABLES: return jsonify({'error': 'Invalid table'}), 400
    data = request.json
    fields = TABLES[table]['fields']
    data['updated_at'] = now_str()
    if 'history' in data and isinstance(data['history'], list):
        data['history'] = json.dumps(data['history'], ensure_ascii=False)
    cols = [f for f in fields if f in data and f != 'id']
    vals = [data[f] for f in cols]
    vals.append(item_id)
    set_clause = ','.join([f'{f}=?' for f in cols])
    conn = get_db()
    conn.execute(f'UPDATE {table} SET {set_clause} WHERE id = ?', vals)
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/<table>/<item_id>', methods=['DELETE'])
def delete(table, item_id):
    if table not in TABLES: return jsonify({'error': 'Invalid table'}), 400
    conn = get_db()
    attachments = conn.execute('SELECT stored_name FROM attachments WHERE parent_table=? AND parent_id=?', (table, item_id)).fetchall()
    for att in attachments:
        filepath = os.path.join(UPLOAD_DIR, att['stored_name'])
        if os.path.exists(filepath): os.remove(filepath)
    conn.execute('DELETE FROM attachments WHERE parent_table=? AND parent_id=?', (table, item_id))
    conn.execute(f'DELETE FROM {table} WHERE id = ?', (item_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/stats', methods=['GET'])
def stats():
    conn = get_db()
    result = {}
    for table in TABLES:
        count = conn.execute(f'SELECT COUNT(*) as cnt FROM {table}').fetchone()
        result[table] = count['cnt']
    conn.close()
    return jsonify(result)

@app.route('/api/upload/<table>/<parent_id>', methods=['POST'])
def upload_file(table, parent_id):
    if table not in TABLES: return jsonify({'error': 'Invalid table'}), 400
    if 'file' not in request.files: return jsonify({'error': 'No file provided'}), 400
    files = request.files.getlist('file')
    results = []
    for f in files:
        if f.filename == '': continue
        original_name = f.filename
        ext = os.path.splitext(original_name)[1].lower()
        if ext not in ALL_ALLOWED:
            results.append({'name': original_name, 'error': f'허용되지 않는 파일 형식: {ext}'})
            continue
        file_id = str(uuid_lib.uuid4())
        stored_name = file_id + ext
        filepath = os.path.join(UPLOAD_DIR, stored_name)
        f.save(filepath)
        file_size = os.path.getsize(filepath)
        file_type = get_file_category(ext)
        mime_type = f.content_type or ''
        conn = get_db()
        conn.execute('''INSERT INTO attachments (id, parent_table, parent_id, original_name, stored_name, file_size, file_type, mime_type, uploaded_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (file_id, table, parent_id, original_name, stored_name, file_size, file_type, mime_type, now_str()))
        conn.commit()
        conn.close()
        results.append({'id': file_id, 'name': original_name, 'size': file_size, 'type': file_type, 'success': True})
    return jsonify({'files': results})

@app.route('/api/attachments/<table>/<parent_id>', methods=['GET'])
def get_attachments(table, parent_id):
    conn = get_db()
    rows = conn.execute('SELECT * FROM attachments WHERE parent_table=? AND parent_id=? ORDER BY uploaded_at DESC', (table, parent_id)).fetchall()
    conn.close()
    return jsonify(rows_to_list(rows))

@app.route('/api/file/<file_id>', methods=['GET'])
def get_file(file_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM attachments WHERE id=?', (file_id,)).fetchone()
    conn.close()
    if row is None: return jsonify({'error': 'File not found'}), 404
    filepath = os.path.join(UPLOAD_DIR, row['stored_name'])
    if not os.path.exists(filepath): return jsonify({'error': 'File missing'}), 404
    return send_file(filepath, download_name=row['original_name'], as_attachment=False)

@app.route('/api/download/<file_id>', methods=['GET'])
def download_file(file_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM attachments WHERE id=?', (file_id,)).fetchone()
    conn.close()
    if row is None: return jsonify({'error': 'File not found'}), 404
    filepath = os.path.join(UPLOAD_DIR, row['stored_name'])
    if not os.path.exists(filepath): return jsonify({'error': 'File missing'}), 404
    return send_file(filepath, download_name=row['original_name'], as_attachment=True)

@app.route('/api/attachment/<file_id>', methods=['DELETE'])
def delete_attachment(file_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM attachments WHERE id=?', (file_id,)).fetchone()
    if row:
        filepath = os.path.join(UPLOAD_DIR, row['stored_name'])
        if os.path.exists(filepath): os.remove(filepath)
        conn.execute('DELETE FROM attachments WHERE id=?', (file_id,))
        conn.commit()
    conn.close()
    return jsonify({'success': True})

# ─── 엑셀 차트 포함 다운로드 API ───────────────────────
@app.route('/api/analytics/excel-chart', methods=['POST'])
def analytics_excel_chart():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter
        import io

        data = request.json
        month_label = data.get('monthLabel', '전체')
        by_type = data.get('byType', [])
        by_line = data.get('byLine', [])
        by_company = data.get('byCompany', [])
        by_team = data.get('byTeam', [])
        line_types_data = data.get('lineTypes', {})
        total = data.get('total', 0)

        wb = Workbook()
        tf = Font(bold=True, size=14, color='FFFFFF')
        tfl = PatternFill('solid', fgColor='1E40AF')
        hf = Font(bold=True, size=11, color='FFFFFF')
        hfl = PatternFill('solid', fgColor='334155')
        bf = Font(bold=True, size=10)
        totf = Font(bold=True, size=11, color='FFFFFF')
        totfl = PatternFill('solid', fgColor='1E40AF')
        evfl = PatternFill('solid', fgColor='F8FAFC')
        brd = Border(left=Side(style='thin',color='E2E8F0'),right=Side(style='thin',color='E2E8F0'),top=Side(style='thin',color='E2E8F0'),bottom=Side(style='thin',color='E2E8F0'))
        ca = Alignment(horizontal='center', vertical='center')
        la = Alignment(horizontal='left', vertical='center')
        colors = ['3B82F6','EF4444','F59E0B','10B981','8B5CF6','EC4899','06B6D4','F97316','6366F1','14B8A6','E11D48','A855F7','0EA5E9','84CC16','D946EF']

        def make_pivot_sheet(ws, title, items, col1):
            ws.merge_cells('A1:C1')
            ws['A1']=title; ws['A1'].font=tf; ws['A1'].fill=tfl; ws['A1'].alignment=ca
            for i,h in enumerate([col1,'건수','비율(%)'],1):
                c=ws.cell(row=2,column=i,value=h); c.font=hf; c.fill=hfl; c.alignment=ca; c.border=brd
            for ri,d in enumerate(items,3):
                ws.cell(row=ri,column=1,value=d['name']).alignment=la
                ws.cell(row=ri,column=2,value=d['count']).font=bf
                ws.cell(row=ri,column=3,value=float(d['rate']))
                if(ri-3)%2==0:
                    for c in range(1,4): ws.cell(row=ri,column=c).fill=evfl
                for c in range(1,4): ws.cell(row=ri,column=c).border=brd; ws.cell(row=ri,column=c).alignment=ca
            tr=len(items)+3
            for i,v in enumerate(['합계',total,100.0],1):
                c=ws.cell(row=tr,column=i,value=v); c.font=totf; c.fill=totfl; c.alignment=ca; c.border=brd
            ws.column_dimensions['A'].width=20; ws.column_dimensions['B'].width=12; ws.column_dimensions['C'].width=12

        # 시트1: 불량유형별
        ws1=wb.active; ws1.title='불량유형별'
        make_pivot_sheet(ws1, f'📊 {month_label} 불량 유형별 집계', by_type, '불량유형')
        if by_type:
            pie=PieChart(); pie.title=f'{month_label} 불량 유형별 분포'; pie.style=10; pie.width=18; pie.height=12
            cats=Reference(ws1,min_col=1,min_row=3,max_row=2+len(by_type))
            vals=Reference(ws1,min_col=2,min_row=2,max_row=2+len(by_type))
            pie.add_data(vals,titles_from_data=True); pie.set_categories(cats)
            pie.dataLabels=DataLabelList(); pie.dataLabels.showCatName=True; pie.dataLabels.showVal=True; pie.dataLabels.showPercent=True; pie.dataLabels.separator='\n'
            for i in range(len(by_type)):
                pt=DataPoint(idx=i); pt.graphicalProperties.solidFill=colors[i%len(colors)]; pie.series[0].data_points.append(pt)
            ws1.add_chart(pie,'E2')

        # 시트2: 호기별
        ws2=wb.create_sheet('호기별')
        make_pivot_sheet(ws2, f'📊 {month_label} 호기별 불량 집계', by_line, '호기')
        if by_line:
            bar=BarChart(); bar.type='col'; bar.title=f'{month_label} 호기별 불량 건수'; bar.style=10; bar.width=18; bar.height=12
            cats=Reference(ws2,min_col=1,min_row=3,max_row=2+len(by_line))
            vals=Reference(ws2,min_col=2,min_row=2,max_row=2+len(by_line))
            bar.add_data(vals,titles_from_data=True); bar.set_categories(cats)
            bar.dataLabels=DataLabelList(); bar.dataLabels.showVal=True
            bar.series[0].graphicalProperties.solidFill='3B82F6'
            ws2.add_chart(bar,'E2')

        # 시트3: 호기별 불량유형 스택형
        if line_types_data:
            ws3=wb.create_sheet('호기별_불량유형')
            all_types=sorted(set(t for types in line_types_data.values() for t in types))
            lines=sorted(line_types_data.keys())
            ws3.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(all_types)+2)
            ws3['A1']=f'📊 {month_label} 호기별 불량유형 현황 (총 {total}건)'; ws3['A1'].font=tf; ws3['A1'].fill=tfl; ws3['A1'].alignment=ca
            hdrs=['호기']+all_types+['합계']
            for i,h in enumerate(hdrs,1):
                c=ws3.cell(row=2,column=i,value=h); c.font=hf; c.fill=hfl; c.alignment=ca; c.border=brd
            for ri,line in enumerate(lines,3):
                ws3.cell(row=ri,column=1,value=line).alignment=la
                rt=0
                for ti,t in enumerate(all_types,2):
                    v=line_types_data.get(line,{}).get(t,0)
                    cell=ws3.cell(row=ri,column=ti,value=v if v else None); cell.alignment=ca
                    if v: cell.font=bf
                    rt+=v
                ws3.cell(row=ri,column=len(all_types)+2,value=rt).font=Font(bold=True,size=10,color='1E40AF'); ws3.cell(row=ri,column=len(all_types)+2).alignment=ca
                if(ri-3)%2==0:
                    for c in range(1,len(hdrs)+1): ws3.cell(row=ri,column=c).fill=evfl
                for c in range(1,len(hdrs)+1): ws3.cell(row=ri,column=c).border=brd
            ws3.column_dimensions['A'].width=12
            for i in range(2,len(hdrs)+1): ws3.column_dimensions[get_column_letter(i)].width=12
            stk=BarChart(); stk.type='col'; stk.grouping='stacked'; stk.title=f'{month_label} 호기별 불량유형 현황'
            stk.style=10; stk.width=24; stk.height=14
            cats=Reference(ws3,min_col=1,min_row=3,max_row=2+len(lines))
            for ti in range(len(all_types)):
                vals=Reference(ws3,min_col=ti+2,min_row=2,max_row=2+len(lines))
                stk.add_data(vals,titles_from_data=True); stk.set_categories(cats)
                stk.series[ti].graphicalProperties.solidFill=colors[ti%len(colors)]
            stk.dataLabels=DataLabelList(); stk.dataLabels.showVal=True; stk.dataLabels.number_format='#'
            ws3.add_chart(stk,f'A{len(lines)+5}')

        # 시트4: 업체별
        ws4=wb.create_sheet('업체별')
        make_pivot_sheet(ws4, f'📊 {month_label} 업체별 불량 집계', by_company, '업체')

        # 시트5: 팀별
        ws5=wb.create_sheet('팀별')
        team_items = [{'name':d['name'] or '미분류','count':d['count'],'rate':d['rate']} for d in by_team]
        make_pivot_sheet(ws5, f'📊 {month_label} 팀별 불량 건수 및 비율', team_items, '팀')
        if by_team:
            pie2=PieChart(); pie2.title=f'{month_label} 팀별 불량 비율'; pie2.style=10; pie2.width=16; pie2.height=11
            cats=Reference(ws5,min_col=1,min_row=3,max_row=2+len(by_team))
            vals=Reference(ws5,min_col=2,min_row=2,max_row=2+len(by_team))
            pie2.add_data(vals,titles_from_data=True); pie2.set_categories(cats)
            pie2.dataLabels=DataLabelList(); pie2.dataLabels.showCatName=True; pie2.dataLabels.showPercent=True; pie2.dataLabels.showVal=True
            for i in range(len(by_team)):
                pt=DataPoint(idx=i); pt.graphicalProperties.solidFill=colors[i%len(colors)]; pie2.series[0].data_points.append(pt)
            ws5.add_chart(pie2,'E2')

        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, download_name=f'불량집계_{month_label}_{datetime.now().strftime("%Y-%m-%d")}.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500
