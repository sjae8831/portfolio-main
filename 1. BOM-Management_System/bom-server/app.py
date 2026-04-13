"""
BOM 배합 관리 시스템 v4 - Flask 서버 (PythonAnywhere 호환)
- 호기별 BOM 지원
- 머신 그룹 필터링
"""
import json, os, shutil
from datetime import datetime
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS

app = Flask(__name__, static_folder="static", static_url_path="/static")
CORS(app)

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE, "data.json")
BACKUP_DIR = os.path.join(BASE, "backups")


def load_data():
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_data(data):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    if os.path.exists(DATA_FILE):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(DATA_FILE, os.path.join(BACKUP_DIR, f"data_{ts}.json"))
        backups = sorted(os.listdir(BACKUP_DIR))
        for old in backups[:-30]:
            os.remove(os.path.join(BACKUP_DIR, old))
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)


@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/api/data")
def get_data():
    return jsonify(load_data())


# ─── Products CRUD ───

@app.route("/api/products", methods=["POST"])
def add_product():
    data = load_data()
    p = request.json
    # seq는 문자열 코드 형식 (예: "1-99")
    existing_seqs = [x["seq"] for x in data["products"]]
    # 자동 seq: prefix-N
    prefix = p.get("machinePrefix", "1")
    max_n = 0
    for s in existing_seqs:
        parts = str(s).split("-")
        if len(parts) == 2 and parts[0] == prefix:
            try:
                n = int(parts[1])
                if n > max_n: max_n = n
            except:
                pass
    p["seq"] = f"{prefix}-{max_n + 1}"
    p.setdefault("materials", [])
    p.setdefault("additives", [])
    p.setdefault("layerRatios", {"inner": 0.333, "mid": 0.334, "outer": 0.333})
    p.setdefault("totalCost", 0)
    p.setdefault("slipInfo", None)
    p.setdefault("machines", [])
    p.setdefault("machineGroup", "")
    data["products"].append(p)
    save_data(data)
    return jsonify({"ok": True, "seq": p["seq"]})


@app.route("/api/products/<seq>", methods=["PUT"])
def update_product(seq):
    data = load_data()
    u = request.json
    for i, p in enumerate(data["products"]):
        if str(p["seq"]) == str(seq):
            u["seq"] = p["seq"]
            data["products"][i] = u
            save_data(data)
            return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "not found"}), 404


@app.route("/api/products/<seq>", methods=["DELETE"])
def delete_product(seq):
    data = load_data()
    data["products"] = [p for p in data["products"] if str(p["seq"]) != str(seq)]
    save_data(data)
    return jsonify({"ok": True})


# ─── BOM CRUD ───

@app.route("/api/bom", methods=["POST"])
def add_bom():
    data = load_data()
    data["bom"].append(request.json)
    data["companies"] = sorted(set(b["company"] for b in data["bom"] if b.get("company")))
    data["productNames"] = sorted(set(b["product"] for b in data["bom"] if b.get("product")))
    save_data(data)
    return jsonify({"ok": True})


@app.route("/api/bom/delete", methods=["POST"])
def delete_bom():
    data = load_data()
    t = request.json
    data["bom"] = [b for b in data["bom"] if not (
        b["company"] == t["company"] and b["product"] == t["product"]
        and b.get("color") == t.get("color") and abs(b["ppm"] - t["ppm"]) < 0.01
        and b.get("machineGroup","") == t.get("machineGroup",""))]
    data["companies"] = sorted(set(b["company"] for b in data["bom"] if b.get("company")))
    data["productNames"] = sorted(set(b["product"] for b in data["bom"] if b.get("product")))
    save_data(data)
    return jsonify({"ok": True})


# ─── Materials CRUD ───

@app.route("/api/materials", methods=["POST"])
def add_material():
    data = load_data()
    data["rawMaterials"].append(request.json)
    save_data(data)
    return jsonify({"ok": True})


@app.route("/api/materials/update", methods=["POST"])
def update_material():
    data = load_data()
    u = request.json
    orig = u.pop("originalName", None)
    mat_type = u.pop("materialType", "")
    mat_maker = u.pop("materialMaker", "")
    for i, m in enumerate(data["rawMaterials"]):
        if m["name"] == orig:
            data["rawMaterials"][i] = u
            if "materialTypes" not in data:
                data["materialTypes"] = {}
            if orig != u["name"] and orig in data["materialTypes"]:
                del data["materialTypes"][orig]
            if mat_type:
                data["materialTypes"][u["name"]] = {"type": mat_type, "maker": mat_maker, "cas": ""}
            elif u["name"] in data["materialTypes"]:
                del data["materialTypes"][u["name"]]
            save_data(data)
            return jsonify({"ok": True})
    return jsonify({"ok": False}), 404


@app.route("/api/materials/delete", methods=["POST"])
def delete_material():
    data = load_data()
    data["rawMaterials"] = [m for m in data["rawMaterials"] if m["name"] != request.json["name"]]
    save_data(data)
    return jsonify({"ok": True})


# ─── Excel Upload ───

@app.route("/api/upload-excel", methods=["POST"])
def upload_excel():
    try:
        import openpyxl
        f = request.files.get("file")
        if not f:
            return jsonify({"ok": False, "error": "파일 없음"}), 400
        wb = openpyxl.load_workbook(f, data_only=True)
        new_data = parse_workbook(wb)
        save_data(new_data)
        return jsonify({"ok": True, "products": len(new_data["products"]), "bom": len(new_data["bom"])})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def sf(val, default=0):
    if val is None: return default
    try: return float(val)
    except: return default


def parse_workbook(wb):
    MACHINES = ['1호기', '2호기', '3호기', '4호기', '5호기', '6호기', '7호기', '8호기', '9호기', '10호기', '11호기']
    PREFIX_INFO = {
        '1':  {'machines': ['1호기', '2호기'], 'layerCount': 1, 'label': '1레이어 (1,2호기)'},
        '2':  {'machines': ['3호기', '7호기'], 'layerCount': 2, 'label': '2레이어 (3,7호기)'},
        '35': {'machines': ['5호기'],           'layerCount': 3, 'label': '3레이어 (5호기)'},
        '3K': {'machines': ['4호기', '6호기', '8호기'], 'layerCount': 3, 'label': '3레이어 국산 (4,6,8호기)'},
        '3R': {'machines': ['10호기', '11호기'], 'layerCount': 3, 'label': '3레이어 라이펜 (10,11호기)'},
        '5':  {'machines': ['9호기'],           'layerCount': 5, 'label': '5레이어 (9호기)'}
    }

    # 원재료비 데이터
    ws3 = wb["원재료비 데이터"]
    rows3 = list(ws3.iter_rows(min_row=1, values_only=True))
    raw = []
    mat_types = {}
    current_type = ""
    for i in range(3, len(rows3)):
        r = rows3[i]
        if r[0] and str(r[0]).strip(): current_type = str(r[0]).strip()
        name = str(r[1]).strip() if r[1] else ""
        if not name: continue
        price = sf(r[2])
        if price <= 0: continue
        raw.append({
            "name": name, "price": price,
            "company": str(r[3]).strip() if r[3] else "",
            "mi": str(r[4]).strip() if r[4] else "",
            "density": str(r[5]).strip() if r[5] else "",
            "slip": sf(r[6]), "ppa": sf(r[7]), "ab": sf(r[8]), "ao": sf(r[9])
        })
        if current_type:
            mat_types[name] = {"type": current_type, "maker": str(r[3]).strip() if r[3] else "", "cas": ""}

    # BOM 작성용
    ws1 = wb["BOM 작성용"]
    rows1 = list(ws1.iter_rows(min_row=1, values_only=True))
    starts = []
    for i in range(4, len(rows1)):
        v = rows1[i][0]
        if v is not None and str(v).strip() not in ("", "순번"):
            starts.append(i)

    products = []
    for idx, start in enumerate(starts):
        r = rows1[start]
        seq_raw = str(r[0]).strip()
        prefix = seq_raw.split("-")[0]
        pinfo = PREFIX_INFO.get(prefix, {"machines": [], "layerCount": 1, "label": ""})
        name = str(r[3]).replace("\n", " / ").strip() if r[3] else ""
        ppm = round(sf(r[1]), 2)
        color = str(r[4]).strip() if r[4] else ""
        active_machines = [MACHINES[j] for j in range(11) if r[5+j] == True]
        li_src = rows1[start-1] if start > 0 else None
        def get_lr(src, col, default):
            if src and len(src) > col and src[col] is not None:
                v = sf(src[col], None)
                if v is not None: return round(v, 4)
            return default
        lc = pinfo["layerCount"]
        li = get_lr(li_src, 18, 1.0 if lc==1 else 0.5 if lc==2 else 0.333)
        lm = get_lr(li_src, 20, 0.0 if lc<=2 else 0.334)
        lo = get_lr(li_src, 22, 0.0 if lc==1 else 0.5 if lc==2 else 0.333)
        materials, additives, in_add, total_cost, slip_info = [], [], False, 0, None
        next_start = starts[idx+1] if idx+1 < len(starts) else start+18
        for j in range(start, min(next_start, start+18, len(rows1))):
            rr = rows1[j]
            mn16 = rr[16] if len(rr) > 16 else None
            if mn16 and "비율 합계" in str(mn16): in_add = True; continue
            if mn16 and len(rr)>37 and rr[37] and "슬립" in str(rr[37]):
                slip_info = {"inner":sf(rr[39] if len(rr)>39 else None),"mid":sf(rr[40] if len(rr)>40 else None),"outer":sf(rr[41] if len(rr)>41 else None),"total":sf(rr[44] if len(rr)>44 else None)}; continue
            if mn16 and str(mn16).strip():
                e = {"name":str(mn16).strip(),"inner":round(sf(rr[18] if len(rr)>18 else None),6),"mid":round(sf(rr[20] if len(rr)>20 else None),6),"outer":round(sf(rr[22] if len(rr)>22 else None),6),"ppm":round(sf(rr[38] if len(rr)>38 else None),2),"ppmInner":round(sf(rr[39] if len(rr)>39 else None),6),"ppmMid":round(sf(rr[40] if len(rr)>40 else None),6),"ppmOuter":round(sf(rr[41] if len(rr)>41 else None),6),"ppmSub":round(sf(rr[44] if len(rr)>44 else None),6),"price":round(sf(rr[34] if len(rr)>34 else None),2),"cost":round(sf(rr[35] if len(rr)>35 else None),2)}
                total_cost += e["cost"]
                e = {k:v for k,v in e.items() if v != 0 or k in ("ppm","name")}
                (additives if in_add else materials).append(e)
        tc = sf(r[36] if len(r)>36 else None, None)
        if tc is not None: total_cost = round(tc, 2)
        additives = [a for a in additives if any(v for k,v in a.items() if k not in ("name","ppm"))]
        products.append({"seq":seq_raw,"name":name,"ppm":round(ppm,2),"color":color,"machines":active_machines,"machineGroup":pinfo["label"],"machinePrefix":prefix,"layerCount":pinfo["layerCount"],"layerRatios":{"inner":li,"mid":lm,"outer":lo},"materials":materials,"additives":additives,"totalCost":round(total_cost,2),"slipInfo":slip_info})

    # BOM 산출
    ws2 = wb["BOM 산출"]
    rows2 = list(ws2.iter_rows(min_row=1, values_only=True))
    bom = []
    header_row = None
    for i, r in enumerate(rows2):
        if r[0] == "코드":
            header_row = i
            break
    if header_row is not None:
        for i in range(header_row+1, len(rows2)):
            r = rows2[i]
            if r[0] and r[3] and r[4]:
                code = str(r[0]).strip()
                prefix = code.split("-")[0]
                pinfo = PREFIX_INFO.get(prefix, {})
                bom.append({
                    "code": code,
                    "hogi": str(r[1]).replace("\n",", ").strip() if r[1] else "",
                    "company": str(r[3]).strip(),
                    "product": str(r[4]).strip(),
                    "color": str(r[5]).strip() if r[5] else "",
                    "ppm": round(sf(r[6]), 2),
                    "bomCost": round(sf(r[7]), 2),
                    "lossRate": round(sf(r[8]), 4),
                    "productivity": round(sf(r[10]), 4),
                    "unitCost": round(sf(r[11]), 2),
                    "machineGroup": pinfo.get("label", ""),
                    "machines": pinfo.get("machines", [])
                })

    return {
        "products": products,
        "bom": bom,
        "rawMaterials": raw,
        "companies": sorted(set(b["company"] for b in bom if b.get("company"))),
        "productNames": sorted(set(b["product"] for b in bom if b.get("product"))),
        "machineGroups": [p["label"] for p in PREFIX_INFO.values()],
        "materialTypes": mat_types
    }


if __name__ == "__main__":
    print("  BOM 배합 관리 시스템 v4 | http://localhost:5000")
    app.run(host="0.0.0.0", port=5000, debug=False)
